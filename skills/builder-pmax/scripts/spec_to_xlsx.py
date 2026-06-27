#!/usr/bin/env python
"""
spec_to_xlsx.py — render a polished, client-ready blueprint workbook from a campaign-spec.json

Usage:
    python spec_to_xlsx.py <campaign-spec.json> --output blueprint.xlsx

Data-driven: one source of truth (the spec). Handles BOTH PMax/Demand Gen (asset_groups) and Search/Branded
(ad_groups). 5 sheet types: Overview (campaign settings + split + summary) · Extensions (campaign-level
sitelinks/callouts/snippets/prices/promo placeholder) · one AG sheet per asset group OR ad group (copy with
GREEN/AMBER/RED char-count coloring; PMax AGs add audience signals + listing group + creative brief; Search
AGs add keywords + ad-group negatives + RSA + paths) · Negative Keywords · Checklist. Business-agnostic;
verifies counts/limits after building.
"""
import argparse
import json
import sys
import unicodedata

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    sys.exit("Missing dependency: openpyxl. Install with:  python -m pip install openpyxl")

NAVY, BLUE = "1B4F72", "2E86C1"
INK, MUT = "1B2733", "5B6B7B"
OK_F, WARN_F, OVER_F = "DCF2E3", "FCEFD6", "FAD9D5"
OK_T, WARN_T, OVER_T = "1E8E5A", "B9770E", "C0392B"
ZEBRA = "F4F8FC"
AG_TABS = ["2E86C1", "F39C12", "8E44AD", "27AE60", "E74C3C", "16A085"]
THIN = Border(*[Side(style="thin", color="D9E2EC")] * 4)
HAIR = Side(style="thin", color="D9E2EC")


def glen(s):
    s = unicodedata.normalize("NFC", str(s or "")).strip()
    return sum(2 if unicodedata.east_asian_width(c) in ("W", "F") else 1 for c in s)


def _yn(v):
    return "ON" if v is True else "OFF" if v is False else "—"


def fmt_bidding(bidding):
    """Human label per strategy — no false 'learning' framing when a target is simply absent."""
    strat = (bidding.get("strategy") or "").lower()
    troas = bidding.get("target_roas")
    if strat == "target_impression_share":
        tis = bidding.get("target_impression_share") or {}
        loc = str(tis.get("location", "")).replace("_", " ")
        pct = tis.get("target_percent")
        cap = tis.get("max_cpc_bid_limit")
        s = f"Target Impression Share · {loc} {pct}%" if pct else "Target Impression Share"
        return s + (f" · max CPC {cap}" if cap else "")
    if strat in ("maximize_conversion_value", "maxconv-value-then-troas"):
        return f"Maximize Conversion Value · tROAS {troas}" if troas else "Maximize Conversion Value (no target set)"
    if strat in ("maximize_conversions", "maxconv-then-tcpa"):
        tcpa = bidding.get("target_cpa")
        return f"Maximize Conversions · tCPA {tcpa}" if tcpa else "Maximize Conversions"
    if strat == "target_roas":
        return f"Target ROAS {troas}" if troas else "Target ROAS"
    if strat == "target_cpa":
        return f"Target CPA {bidding.get('target_cpa')}"
    return bidding.get("strategy") or "—"


def sheet_name(s, used):
    n = "".join(c for c in str(s) if c not in '[]:*?/\\')[:31] or "Sheet"
    base, i = n, 1
    while n in used:
        n = f"{base[:28]}_{i}"; i += 1
    used.add(n)
    return n


def banner(ws, span, title, sub=""):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=16, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    if sub:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
        s = ws.cell(row=2, column=2 - 1, value=sub)
        s.font = Font(italic=True, size=10, color=MUT); s.alignment = Alignment(indent=1)
    for col in range(1, span + 1):
        ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor=NAVY)


def header(ws, row, values):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def row_cells(ws, row, values, zebra=False, counts=None):
    """counts: dict {col_index: limit} to colour the count cell green/amber/red."""
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN
        cell.font = Font(size=10, color=INK, bold=(c == 1))
        if zebra and row % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=ZEBRA)
        if counts and c in counts and isinstance(v, int):
            limit = counts[c]
            f, t = (OVER_F, OVER_T) if v > limit else (WARN_F, WARN_T) if v >= limit * 0.9 else (OK_F, OK_T)
            cell.fill = PatternFill("solid", fgColor=f); cell.font = Font(size=10, bold=True, color=t)
    return row + 1


def section(ws, row, title, span):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(bold=True, size=11, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="5499C7")
    c.alignment = Alignment(indent=1, vertical="center")
    ws.row_dimensions[row].height = 20
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="5499C7")
    return row + 1


def metric_cards(ws, row, cards):
    """cards: list of (label, value). Renders label row + value row as colored boxes."""
    for i, (label, _) in enumerate(cards):
        col = 1 + i * 2
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        c = ws.cell(row=row, column=col, value=label.upper())
        c.font = Font(size=9, bold=True, color=MUT); c.fill = PatternFill("solid", fgColor="EAF2FB")
        c.alignment = Alignment(horizontal="center")
    for i, (_, val) in enumerate(cards):
        col = 1 + i * 2
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        c = ws.cell(row=row + 1, column=col, value=val)
        c.font = Font(size=15, bold=True, color=NAVY); c.fill = PatternFill("solid", fgColor="F4F9FD")
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row + 1].height = 26
    return row + 3


def flatten_listing(node, path, rows):
    if not isinstance(node, dict):
        return
    if node.get("type") == "subdivision":
        p = path + [f"{node.get('dimension')}={node.get('value')}"]
        for ch in node.get("children") or []:
            flatten_listing(ch, p, rows)
    else:
        dim, val = node.get("dimension"), node.get("value")
        filt = "Everything else" if (dim == "other" or val in (None, "", "other")) else f"{dim}={val}"
        rows.append((" > ".join(path) or "All", filt, (node.get("action") or "").upper()))


def build(spec, out):
    wb = openpyxl.Workbook()
    used = set()
    camp = spec.get("campaign", {})
    ctype = spec.get("campaign_type", "")
    is_search = ctype in ("search", "branded_search")
    is_dg = ctype == "demand_gen"   # Demand Gen has no campaign-level extensions (creative/audience-led)
    ags = spec.get("asset_groups") or []
    adgs = spec.get("ad_groups") or []
    units = adgs if is_search else ags
    unit_label = "Ad groups" if is_search else "Asset groups"
    ext = spec.get("extensions") or {}
    bidding = camp.get("bidding", {})
    cur = spec.get("account", {}).get("currency", "")
    bid_phase = fmt_bidding(bidding)
    kind = "Search" if is_search else "Performance Max"

    # ================= SHEET 1: OVERVIEW (campaign-level settings) =================
    ws = wb.active; ws.title = sheet_name("Overview", used)
    ws.sheet_properties.tabColor = NAVY
    banner(ws, 6, f"{kind} — {camp.get('name','(unnamed)')}",
           "Campaign blueprint · created PAUSED · review → enable")
    r = 4
    r = metric_cards(ws, r, [
        ("Daily budget", f"{camp.get('daily_budget','')} {cur}"),
        (unit_label, str(len(units))),
        ("Geo", ", ".join(camp.get("geo_targeting") or []) or "—"),
    ])
    r = section(ws, r, "Campaign settings", 6)
    settings = [("Bid strategy", bid_phase),
                ("Conversion goal", ", ".join((camp.get("conversion_goals") or {}).get("goals") or [])),
                ("Conversion goal scope", (camp.get("conversion_goals") or {}).get("scope"))]
    if is_search:
        nets = camp.get("networks") or {}
        settings.append(("Networks", f"Search {_yn(nets.get('search', True))} · "
                         f"Search Partners {_yn(nets.get('search_partners'))} · "
                         f"Display expansion {_yn(nets.get('display_expansion'))}"))
    else:
        be = camp.get("brand_exclusion", {})
        settings += [
            ("Brand exclusion", _yn(be.get("enabled"))
             + (("  [" + ", ".join(be.get("brand_list_terms") or []) + "]") if be.get("brand_list_terms") else "")),
            ("Final URL expansion", _yn(camp.get("final_url_expansion"))),
            ("Auto-created assets", _yn(camp.get("auto_created_assets"))),
        ]
    settings.append(("Languages", ", ".join(camp.get("language_targeting") or []) or "—"))
    for k, v in settings:
        r = row_cells(ws, r, [k, v])
    r += 1
    r = section(ws, r, "Split strategy", 6)
    ss = spec.get("split_strategy", {})
    r = row_cells(ws, r, ["Campaign axis", ss.get("campaign_axis", "")])
    r = row_cells(ws, r, ["Asset-group axis", ss.get("asset_group_axis", "")])
    r = row_cells(ws, r, ["Rationale", ss.get("rationale", "")])
    r += 1
    if is_search:
        r = section(ws, r, "Ad groups (detail in each AG sheet)", 6)
        header(ws, r, ["#", "Ad group", "Final URL", "Keywords", "Headlines", "Negatives"]); r += 1
        for i, ag in enumerate(adgs, 1):
            rsa = ag.get("rsa") or {}
            r = row_cells(ws, r, [i, ag.get("name"), rsa.get("final_url"),
                                  len(ag.get("keywords") or []), len(rsa.get("headlines") or []),
                                  len(ag.get("negatives") or [])], zebra=True)
    else:
        r = section(ws, r, "Asset groups (detail in each AG sheet)", 6)
        header(ws, r, ["#", "Asset group", "Final URL", "Headlines", "Themes", "Aud. signals"]); r += 1
        for i, ag in enumerate(ags, 1):
            r = row_cells(ws, r, [i, ag.get("name"), ag.get("final_url"),
                                  len(ag.get("headlines") or []), len(ag.get("search_themes") or []),
                                  sum(len(v or []) for v in (ag.get("audience_signal") or {}).values())], zebra=True)
    r += 1
    if is_dg:
        r = section(ws, r, "Creative & audiences are the levers (see each AG sheet)", 6)
        r = row_cells(ws, r, ["Note", "Demand Gen has no sitelink/callout/snippet extensions — the creative "
                              "groups + audience signals do the work."])
        r = row_cells(ws, r, ["Campaign negatives", len(spec.get("campaign_negatives") or [])])
    else:
        r = section(ws, r, "Campaign-level assets (see Extensions / Negatives sheets)", 6)
        r = row_cells(ws, r, ["Sitelinks", len(ext.get("sitelinks") or [])])
        r = row_cells(ws, r, ["Callouts", len(ext.get("callouts") or [])])
        r = row_cells(ws, r, ["Structured snippets", len(ext.get("structured_snippets") or [])])
        r = row_cells(ws, r, ["Campaign negatives", len(spec.get("campaign_negatives") or [])])
    for col, wd in zip("ABCDEF", (20, 26, 40, 11, 10, 12)):
        ws.column_dimensions[col].width = wd
    ws.freeze_panes = "A3"

    # ================= SHEET 2: EXTENSIONS (campaign-level) — skip for Demand Gen =================
    if not is_dg:
        ws = wb.create_sheet(sheet_name("Extensions", used)); ws.sheet_properties.tabColor = "117A65"
        banner(ws, 5, "Extensions — CAMPAIGN level",
               "Sitelinks · callouts · structured snippets attach at the campaign level")
        r = 3
        r = section(ws, r, "Sitelinks (text ≤25 · desc ≤35)", 5)
        header(ws, r, ["Text", "Chars", "Desc 1", "Desc 2", "Final URL"]); r += 1
        for sl in ext.get("sitelinks") or []:
            r = row_cells(ws, r, [sl.get("text"), glen(sl.get("text", "")), sl.get("description1"),
                                  sl.get("description2"), sl.get("final_url")], zebra=True, counts={2: 25})
        r += 1
        r = section(ws, r, "Callouts (≤25 each · add 4-10)", 5)
        header(ws, r, ["Callout", "Chars", "", "", ""]); r += 1
        for c in ext.get("callouts") or []:
            r = row_cells(ws, r, [c, glen(c), "", "", ""], counts={2: 25})
        r += 1
        r = section(ws, r, "Structured snippets (header from Google list · ≥3 values · ≤25 each · 'Brands' only for multi-brand)", 5)
        header(ws, r, ["Header", "Value (one per row)", "Chars", "", ""]); r += 1
        for sn in ext.get("structured_snippets") or []:
            vals = sn.get("values") or []
            for k, v in enumerate(vals):
                r = row_cells(ws, r, [sn.get("header") if k == 0 else "", v, glen(v), "", ""],
                              zebra=True, counts={3: 25})
        r += 1
        # Prices — render the real items as a table (one item per row, each field its own column)
        prices = ext.get("prices") or []
        if prices:
            for pa in prices:
                cy = pa.get("currency", "")
                ttl = f"Price asset — {pa.get('type','')} · {pa.get('price_qualifier','')} · {cy}".strip(" ·")
                r = section(ws, r, ttl, 5)
                header(ws, r, ["Item (≤25)", "Description (≤25)", "Price", "Final URL", ""]); r += 1
                for it in pa.get("items") or []:
                    pr = it.get("price", "")
                    pr_disp = f"{pr} {cy}".strip()
                    r = row_cells(ws, r, [it.get("header"), it.get("description"), pr_disp,
                                          it.get("final_url"), ""], zebra=True)
        else:
            r = section(ws, r, "Prices — PLACEHOLDER", 5)
            r = row_cells(ws, r, ["Prices", ext.get("_prices_placeholder")
                                  or "Optional: 3-8 price items {header, description, price, final_url}.",
                                  "", "", ""])
        # Promotions — placeholder reminder
        r = section(ws, r, "Promotions — PLACEHOLDER (add when a sale is live)", 5)
        r = row_cells(ws, r, ["Promotions", f"{len(ext.get('promotions') or [])} set — "
                              + (ext.get("_promotions_placeholder") or "Add when running a sale."), "", "", ""])
        for col, wd in zip("ABCDE", (24, 24, 14, 40, 8)):
            ws.column_dimensions[col].width = wd
        ws.freeze_panes = "A3"

    # ================= SHEET(S) 3..: PER AD GROUP (Search/Branded) =================
    wrap = {"exact": lambda t: f"[{t}]", "phrase": lambda t: f'"{t}"', "broad": lambda t: f"{t}"}
    if is_search:
        for i, ag in enumerate(adgs):
            rsa = ag.get("rsa") or {}
            ws = wb.create_sheet(sheet_name(f"AG{i+1} {ag.get('name','')}", used))
            ws.sheet_properties.tabColor = AG_TABS[i % len(AG_TABS)]
            banner(ws, 4, f"Ad group {i+1}: {ag.get('name','')}", rsa.get("final_url", ""))
            r = 4
            r = row_cells(ws, r, ["Theme", ag.get("theme", "—"), "Final URL", rsa.get("final_url", "")])
            # Keywords (ready-to-copy wrapped) + provenance (where the keyword came from)
            r = section(ws, r, "KEYWORDS (harvested — ready-to-copy)", 4)
            header(ws, r, ["KW", "Keyword", "Match", "Source"]); r += 1
            for j, kw in enumerate(ag.get("keywords") or [], 1):
                mt = (kw.get("match_type") or "broad").lower()
                r = row_cells(ws, r, [f"KW{j}", wrap.get(mt, wrap["broad"])(kw.get("text")), mt,
                                      kw.get("source", "")])
            # Ad-group negatives
            r = section(ws, r, "AD-GROUP NEGATIVES", 4)
            header(ws, r, ["NEG", "Keyword", "Match", ""]); r += 1
            for j, n in enumerate(ag.get("negatives") or [], 1):
                mt = (n.get("match_type") or "phrase").lower()
                r = row_cells(ws, r, [f"NEG{j}", wrap.get(mt, wrap["broad"])(n.get("text")), mt, ""])
            # RSA copy
            for label, key, limit, prefix in [
                ("RSA HEADLINES (≤30 · max 15)", "headlines", 30, "H"),
                ("RSA DESCRIPTIONS (≤90 · max 4)", "descriptions", 90, "D"),
            ]:
                r = section(ws, r, label, 4)
                header(ws, r, ["ID", "Text", "Chars", "Pinned"]); r += 1
                for j, it in enumerate(rsa.get(key) or [], 1):
                    t = it.get("text", "")
                    r = row_cells(ws, r, [f"{prefix}{j}", t, glen(t), it.get("pinned") or ""],
                                  counts={3: limit})
            r = section(ws, r, "DISPLAY PATHS (≤15 each)", 4)
            header(ws, r, ["Path", "Text", "Chars", ""]); r += 1
            for j, p in enumerate(rsa.get("paths") or [], 1):
                r = row_cells(ws, r, [f"P{j}", p, glen(p), ""], counts={3: 15})
            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 60
            ws.column_dimensions["C"].width = 8
            ws.column_dimensions["D"].width = 10
            ws.freeze_panes = "A4"

    # ================= SHEET(S) 3..: PER ASSET GROUP (PMax/Demand Gen) =================
    for i, ag in enumerate([] if is_search else ags):
        ws = wb.create_sheet(sheet_name(f"AG{i+1} {ag.get('name','')}", used))
        ws.sheet_properties.tabColor = AG_TABS[i % len(AG_TABS)]
        banner(ws, 4, f"Asset group {i+1}: {ag.get('name','')}", ag.get("final_url", ""))
        r = 4
        r = row_cells(ws, r, ["Display paths", " / ".join(ag.get("display_paths") or []), "", ""])
        r = row_cells(ws, r, ["Business name", ag.get("business_name", "—"),
                              "Call to action", ag.get("call_to_action", "—")])
        for label, key, limit, prefix in [
            ("HEADLINES (≤30)", "headlines", 30, "H"),
            ("LONG HEADLINES (≤90)", "long_headlines", 90, "LH"),
            ("DESCRIPTIONS (≤90; #1 ≤60)", "descriptions", 90, "D"),
            ("SEARCH THEMES (up to 50)", "search_themes", None, "ST"),
        ]:
            r = section(ws, r, label, 4)
            header(ws, r, ["ID", "Text", "Chars", ""]); r += 1
            for j, t in enumerate(ag.get(key) or [], 1):
                cnt = glen(t)
                lim = 60 if (key == "descriptions" and j == 1) else limit
                r = row_cells(ws, r, [f"{prefix}{j}", t, cnt, ""], counts=({3: lim} if lim else None))
        # Audience signals (AG-level)
        r = section(ws, r, "AUDIENCE SIGNALS (AG-level hints, not hard targeting)", 4)
        header(ws, r, ["AS#", "Signal type", "Value", ""]); r += 1
        n_as = 1
        for stype, vals in (ag.get("audience_signal") or {}).items():
            for v in vals or []:
                r = row_cells(ws, r, [f"AS{n_as}", stype, v, ""]); n_as += 1
        # Listing group (AG-level)
        r = section(ws, r, "LISTING GROUP (product scoping — Everything-Else excluded)", 4)
        header(ws, r, ["LG", "Path", "Filter", "Action"]); r += 1
        rows = []
        flatten_listing(ag.get("listing_group"), [], rows)
        for n_lg, (path, filt, action) in enumerate(rows, 1):
            r = row_cells(ws, r, [f"LG{n_lg}", path, filt, action])
        # Creative brief (AG-level images + video)
        r = section(ws, r, "CREATIVE BRIEF (images + video for this AG)", 4)
        header(ws, r, ["Kind", "Spec / count", "Brief", ""]); r += 1
        for im in ag.get("image_assets") or []:
            cnt = f"{im.get('spec','')} ×{im.get('count', 1)}"
            r = row_cells(ws, r, [f"IMG/{im.get('role')}", cnt, im.get("brief") or im.get("source", ""), ""])
        for vd in ag.get("video_assets") or []:
            r = row_cells(ws, r, [f"VID/{vd.get('orientation')}", vd.get("spec", ""),
                                  vd.get("brief") or vd.get("source", ""), ""])
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 60
        ws.column_dimensions["D"].width = 14
        ws.freeze_panes = "A4"

    # ================= SHEET: NEGATIVE KEYWORDS (campaign-level) =================
    ws = wb.create_sheet(sheet_name("Negative Keywords", used)); ws.sheet_properties.tabColor = "922B21"
    banner(ws, 2, "Campaign negative keywords", "Ready-to-copy: [exact] / \"phrase\" / broad")
    header(ws, 3, ["Keyword (ready-to-copy)", "Match type"]); r = 4
    wrap = {"exact": lambda t: f"[{t}]", "phrase": lambda t: f'"{t}"', "broad": lambda t: f"{t}"}
    for n in spec.get("campaign_negatives") or []:
        mt = (n.get("match_type") or "broad").lower()
        r = row_cells(ws, r, [wrap.get(mt, wrap["broad"])(n.get("text")), mt], zebra=True)
    for sl in spec.get("shared_negative_lists") or []:
        r = row_cells(ws, r, [sl, "(shared list)"], zebra=True)
    ws.column_dimensions["A"].width = 44; ws.column_dimensions["B"].width = 16
    ws.freeze_panes = "A4"

    # ================= SHEET: CHECKLIST =================
    ws = wb.create_sheet(sheet_name("Checklist", used)); ws.sheet_properties.tabColor = "B7950B"
    banner(ws, 3, "Build & QA checklist", "Tick before enabling the campaign")
    header(ws, 3, ["✓", "Step", "Note"]); r = 4
    ext_note = f"{len(ext.get('sitelinks') or [])} sitelinks · {len(ext.get('callouts') or [])} callouts · {len(ext.get('structured_snippets') or [])} snippets"
    common_head = [
        ("Create campaign PAUSED", f"status = {camp.get('status')}"),
        ("Daily budget set", f"{camp.get('daily_budget')} {cur} — confirm vs spend cap"),
        ("Bid strategy", bid_phase),
        ("Conversion goal at CAMPAIGN scope", (camp.get("conversion_goals") or {}).get("scope")),
    ]
    common_tail = ([] if is_dg else [
        ("Extensions added at CAMPAIGN level", ext_note),
        ("Promotion asset added if a sale is running", "placeholder on Extensions sheet"),
    ]) + [
        ("Campaign negatives + shared lists applied", f"{len(spec.get('campaign_negatives') or [])} negatives"),
        ("Reviewed, then ENABLE", "hand monitoring to tracker"),
    ]
    if is_search:
        miss_url = [ag.get("name") for ag in adgs if not (ag.get("rsa") or {}).get("final_url")]
        checks = common_head + [
            ("Every RSA Final URL is LIVE (200, not redirect/404)", "FLAG: " + ", ".join(miss_url) if miss_url else "all ad groups have a final_url"),
            ("Keywords Exact/Phrase only (no broad on brand)", f"{sum(len(a.get('keywords') or []) for a in adgs)} keywords"),
            ("RSA within counts + char limits (≤15 H / ≤4 D)", "see each AG sheet colors"),
            ("NO competitor trademark in any ad text", f"trademark_avoid: {len(spec.get('trademark_avoid') or [])} terms enforced"),
            ("Brand terms negated in conquesting ad groups", "no cross-serve between defense/conquest"),
            ("Brand-exclusion coordinated with PMax/Shopping", "PMax excludes brand so this campaign captures it"),
        ] + common_tail
    else:
        miss_url = [ag.get("name") for ag in ags if not ag.get("final_url")]
        n_imgs = sum(len(ag.get("image_assets") or []) for ag in ags)
        checks = common_head + [
            ("Brand exclusion ON", _yn(camp.get("brand_exclusion", {}).get("enabled"))),
            ("FUE / auto-assets OFF", f"FUE {_yn(camp.get('final_url_expansion'))} · auto {_yn(camp.get('auto_created_assets'))}"),
            ("Every Final URL is LIVE (200, not redirect/404)", "FLAG: " + ", ".join(miss_url) if miss_url else "all AGs have a final_url"),
            ("Listing groups: Everything-Else = EXCLUDED at each level", "no product in two AGs"),
            ("Headlines/long/desc/themes within counts + char limits", "see each AG sheet colors"),
            ("Price asset items live + real prices", f"{sum(len(p.get('items') or []) for p in ext.get('prices') or [])} price items" if ext.get("prices") else "none — consider adding"),
            ("Images uploaded per AG brief", f"{n_imgs} image specs briefed — provide real product shots"),
            ("Audience signals attached per AG", "see each AG sheet"),
        ] + common_tail
    for step, note in checks:
        r = row_cells(ws, r, ["☐", step, note], zebra=True)
    for col, wd in zip("ABC", (5, 52, 50)):
        ws.column_dimensions[col].width = wd
    ws.freeze_panes = "A4"

    wb.save(out)


def verify(out):
    wb = openpyxl.load_workbook(out)
    issues = []
    for name in wb.sheetnames:
        if not name.startswith("AG"):
            continue
        ws = wb[name]
        counts = {"H": 0, "LH": 0, "D": 0, "ST": 0}
        for row in ws.iter_rows():
            vid = str(row[0].value or "")
            for pre in ("LH", "ST", "H", "D"):
                if vid.startswith(pre) and vid[len(pre):].isdigit():
                    counts[pre] += 1
                    break
        for pre, mx in (("H", 15), ("LH", 5), ("D", 5), ("ST", 50)):
            if counts[pre] > mx:
                issues.append(f"{name}: {counts[pre]} {pre} (max {mx})")
    print(f"\n=== blueprint verify: {out} ===")
    print(f"Sheets: {len(wb.sheetnames)} -> {wb.sheetnames}")
    for i in issues:
        print(f"  ! {i}")
    print("Status:", "FAIL" if issues else "PASS")
    return not issues


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    ap = argparse.ArgumentParser()
    ap.add_argument("spec")
    ap.add_argument("--output", required=True)
    args = ap.parse_args()
    with open(args.spec, encoding="utf-8") as f:
        spec = json.load(f)
    if spec.get("campaign_type") not in ("performance_max", "demand_gen", "search", "branded_search"):
        print(f"Note: unknown campaign_type '{spec.get('campaign_type')}' — rendering best-effort.")
    build(spec, args.output)
    verify(args.output)
    print(f"Wrote {args.output}")
